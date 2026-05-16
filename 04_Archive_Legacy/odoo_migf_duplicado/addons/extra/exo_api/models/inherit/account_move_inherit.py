import time
import base64
import openpyxl
import io
import logging
import json
from datetime import datetime, timedelta

from odoo import _, models, fields, http, api
from odoo.http import request
from odoo.exceptions import ValidationError

_logger = logging.getLogger(__name__)

class AccountMoveInherit(models.Model):
    _inherit = "account.move"

    is_automatic_invoice = fields.Boolean('Es una factura automatica?', default=False, required=True, tracking=True)
    lock_invoice = fields.Boolean('Evitar insertar mas cargas en este borrador?', default=False, help="Este campo evita que pueda ser insertado mas cargas en esta factura", tracking=True)
    load_invoice_code = fields.Char('Bloque a la cual pertenece la factura según su agrupación.', tracking=True)
    load_invoice_date = fields.Datetime(string='Fecha de la factura en EXO', tracking=True)
    exo_invoice_list_id = fields.Char('Id del Invoice List En EXO', tracking=True)
    exo_invoice_sequence = fields.Char('Secuencia del Invoice List En EXO', tracking=True)
    last_payment = fields.Date('Último Pago', compute='_compute_payments')
    payment_time = fields.Char('Tiempo de Pago', compute='_compute_payments')
    due_payment_time_days = fields.Float('Dias vencidos realizado el pago', compute='_compute_payments')
    draft_name = fields.Char("Nombre en Borrador", compute="_compute_draft_name")
    block_date_start = fields.Datetime(string='Fecha Inicio del Bloque', tracking=True)
    block_date_end = fields.Datetime(string='Fecha Fin del Bloque', tracking=True)
    is_block_finished = fields.Boolean('La fecha del bloque ha pasado?', readonly=True, compute="_compute_is_block_finished")
    templates_uploaded = fields.Boolean('Plantilla fue cargada?', default=False, help="Indica si ya fueron generados los excel y pdf de esta plantilla", tracking=True)
    subtotal_internal_loads = fields.Float(string="subtotal de Cargas", help='Obtiene el total de las cargas', compute="_compute_subtotal_loads")
    subtotal_internal_loads_additional = fields.Float(string="Subtotal de Adicionales a Cargas", help='Obtiene el total de las cargas', compute="_compute_subtotal_loads")
    subtotal_internal_loads_discount = fields.Float(string="Subtotal de Descuentos por cargas", help='Obtiene el total de las cargas', compute="_compute_subtotal_loads")
    subtotal_internal_loads_subtotal = fields.Float(string="Subtotal de cargas (loads + additionals - discounts)", help='Obtiene el total de las cargas', compute="_compute_subtotal_loads")
    subtotal_internal_discounts = fields.Float(string="subtotal de descuentos internos", help='Obtiene el total de descuentos basado en las cargas', compute="_compute_subtotal_loads")
    automatic_has_loads = fields.Boolean('La Factura Tiene Cargas?', compute="_compute_automatic_no_loads", store=True, help="Indica si la factura fue creada de manera automatica, y si esta tiene cargas registradas...", tracking=True)
    excel_attachment_file = fields.Binary(string='Excel para cargar')
    exo_related = fields.Boolean(string='Related to EXO', default=False, tracking=True)
    discounts_totals = fields.Binary(string="Descuentos Totals", compute='_compute_subtotal_loads')
    provider_checked = fields.Boolean(string="Revisado por el proveedor?", help="Indica si el proveedor ya colocó que fue aprobado o rechazado.", default=False)
    provider_state = fields.Selection([('created', 'Creada'), ('pending', 'Liberado para ser aprobado'), ('agree', 'De acuerdo'), ('disagree', 'En desacuerdo')], string="Estado colocado por el proveedor", default='created')
    disagree_reason = fields.Text("Razon de rechazo")
    

    def print_info(self, obj):
        ("__________________ printing info")
        (obj)

    def move_next(self, value, default_value):
        return next(value, default_value)

    def action_post(self):
        res = super().action_post()
        self.generate_transporter_loads()
        return res

    def generate_transporter_loads(self):
        shipper_line_loads = request.env['account.line.load'].sudo().search([
            ('account_move_id', '=', self.id), ('move_type', '=', 'invoice')
        ])
        shipper_load_numbers = [line_load.original_load_id for line_load in shipper_line_loads] + [line_load.load_id for line_load in shipper_line_loads] if shipper_line_loads else []
        programming_account_line_loads = request.env['account.line.load'].sudo().search([
            ('move_type', '=', 'bill'), ('is_programming', '=', True), ('is_schedule_executed', '=', False),
            '|', ('load_number', 'in', shipper_load_numbers), ('original_load_id', 'in', shipper_load_numbers)
        ])
        programming_account_line_loads.create_in_invoice()

    @api.depends("invoice_line_ids.load_ids")
    def _compute_automatic_no_loads(self):
        for record in self:
            record.automatic_has_loads = len(record.mapped('invoice_line_ids.load_ids')) > 0

    def remove_account_line_load(self):
        for record in self:
            line_loads = self.env['account.line.load'].sudo().search([('account_move_id', '=', record.id)])
            if len(line_loads) > 500:
                raise ValidationError("No se puede eliminar estas cargas.")
            line_loads.sudo().unlink()
        return {'type': 'ir.actions.client', 'tag': 'reload'}

    def insert_line_load_from_excel(self, fields=["No. Carga", 'Load Number', 'CARGA']):
        move_load_numbers_values = self.get_load_numbers_from_excel(fields)
        inserted = {}
        self.insert_line_loads(move_load_numbers_values, inserted)

    def insert_line_load_from_excel_by_orders(self):
        self.insert_line_load_from_excel(['Número de Orden'])

    def insert_line_load_from_lines(self):
        move_load_numbers_values = self.get_loads_numbers_from_line()
        inserted = {}
        self.insert_line_loads(move_load_numbers_values, inserted)

    def insert_line_loads(self, move_load_numbers_values, inserted):
        for move_loads in move_load_numbers_values:
            first_automatic_line = move_loads.get('invoice').line_ids.filtered(lambda l: l.is_automatic_line)
            first_automatic_line = first_automatic_line[0] if first_automatic_line else None
            if first_automatic_line:
                loads_to_insert = []
                for load in move_loads.get('load_numbers_values'):
                    if not inserted.get(load.get('load_number'), False):
                        load_to_insert = (0, 0, {
                            'move_type': 'invoice' if move_loads.get('invoice').move_type == 'out_invoice' else 'bill',
                            'load_number': load.get('load_number'),
                            'account_line_id': load.get('line_id').id if load.get('line_id') else None,
                            'was_restored': True
                        })
                        inserted[load.get('load_number')] = True
                        if load.get('line_id'):
                            load.get('line_id').sudo().write({'load_ids': [load_to_insert]})
                            continue
                        loads_to_insert.append(load_to_insert)
                if loads_to_insert:
                    first_automatic_line.sudo().write({'load_ids': loads_to_insert})

    def get_loads_numbers_from_line(self):
        move_load_numbers_values = []
        for invoice in self:
            load_numbers_values = []
            for line in invoice.line_ids:
                if line and line.name:
                    elements = line.name.split(' - ')
                    if len(elements) >= 2:
                        new_line = elements[1].strip()
                        if new_line:
                            load_numbers_values.append({'load_number': new_line, 'line_id': line})
            move_load_numbers_values.append({'invoice': invoice, 'load_numbers_values': load_numbers_values})
        return move_load_numbers_values

    def get_files(self, record):
        if record.excel_attachment_file:
            return [base64.b64decode(record.excel_attachment_file)]
        files = []
        for attachment in record.attachment_ids:
            if attachment.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                data = base64.b64decode(attachment.datas)
                files.append(data)
        return files

    def get_load_numbers_from_excel(self, fields=["No. Carga", 'Load Number', 'CARGA']):
        move_load_numbers_values = []
        for invoice in self:
            load_numbers_values = []
            files = self.get_files(invoice)
            for data in files:
                wb = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value in fields:
                                column_index = cell.column
                                for data_cell in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
                                    for cell_value in data_cell:
                                        if cell_value.value and str(cell_value.value).strip() and cell_value.value not in fields:
                                            load_numbers_values.append({'load_number': str(cell_value.value).strip(), 'line_id': None})
            if load_numbers_values:
                move_load_numbers_values.append({'invoice': invoice, 'load_numbers_values': load_numbers_values})
        return move_load_numbers_values

    def get_number_orders(self, sheet):
        orders = []
        for row in sheet.iter_rows():
            cell_col_a = row[0] if row[0].column == 1 else None
            for cell in row:
                if cell.value in ['Número de Orden']:
                    column_index = cell.column
                    for data_cell in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
                        for cell_value in data_cell:
                            if cell_value.value and str(cell_value.value).strip() and cell_value.value not in ["Número de Orden"]:
                                if cell_col_a:
                                    col_a_color = cell_col_a.fill.start_color.index
                                    if not col_a_color.startswith('00'):
                                        orders.append(str(cell_value.value).strip())
        return orders

    def find_sheet_with_grand_total(self, invoice):
        files = self.get_files(invoice)
        for data in files:
            wb = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
            for sheet in wb.worksheets:
                for cell in sheet['A']:
                    if cell.value and "Grand Total" in str(cell.value):
                        return {'main_sheet': sheet, 'worksheets': wb.worksheets}
        return None

    def get_load_numbers_by_order_id_from_excel(self):
        move_load_numbers_values = []
        for invoice in self:
            load_numbers_values = []
            main_sheet = self.find_sheet_with_grand_total(invoice)
            orders = []
            if main_sheet:
                orders = self.get_number_orders(main_sheet.get('main_sheet'))
                for sheet in main_sheet.get('worksheets'):
                    if sheet.title == 'Sheet1' and 'PART' not in sheet.title:
                        headers = {}
                        for row in sheet.iter_rows(min_row=1, max_row=1):
                            for cell in row:
                                if cell.value in ["No. Carga", 'Load Number', 'CARGA']:
                                    headers['load_number'] = cell.column
                                elif cell.value == "Número de Orden":
                                    headers['order_number'] = cell.column
                        if not headers:
                            continue
                        for row in sheet.iter_rows(min_row=2):
                            order_number = str(row[headers['order_number'] - 1].value).strip()
                            if order_number in orders:
                                load_number_value = row[headers['load_number'] - 1].value
                                if load_number_value:
                                    load_numbers_values.append({
                                        'load_number': load_number_value.strip(),
                                        'line_id': None,
                                        'order_number': order_number
                                    })
            if load_numbers_values:
                move_load_numbers_values.append({
                    'invoice': invoice,
                    'load_numbers_values': load_numbers_values
                })
        return move_load_numbers_values

    def unlink(self):
        load_ids = self.mapped('invoice_line_ids.load_ids')
        self.env['account.line.deleted.load'].create_deleted(load_ids)
        return super(AccountMoveInherit, self).unlink()

    def generate_partner_bank_fee(self):
        if http.request.env['ir.config_parameter'].sudo().get_param('DIFFERENCE_BANK_FEE') != 'True':
            return
        product_difference_bank_fee = self.env['product.template'].sudo().search([('barcode', '=', 'DIFFERENCE_BANK_FEE')], limit=1)
        if not product_difference_bank_fee:
            raise ValidationError("El Producto Diferencia Bancaria no fue encontrado, revise si existe un producto con el codigo de barra (DIFFERENCE_BANK_FEE)")
        bank_bic_or_swift = http.request.env['ir.config_parameter'].sudo().get_param('PAYMENT_BANK_BIC')
        if not bank_bic_or_swift:
            raise ValidationError("El BANCO ORIGEN no fue encontrado en los parametros, Solicite a su administrador agregar el parametro del banco con el que se les paga a los transportistas. Probablemente el Swift BCBHDOSDXXX del BHD")
        for record in self:
            if record.state == 'draft' and len(record.partner_id.bank_ids) == 1 and record.partner_id.bank_ids.bic == bank_bic_or_swift:
                line_difference_bank_fee = record.invoice_line_ids.filtered(lambda line: line.line_type == 'difference_bank_fee')
                if not line_difference_bank_fee:
                    record.sudo().write({
                        'invoice_line_ids': [(0, 0, {
                            'product_id': product_difference_bank_fee.id,
                            'quantity': 1,
                            'name': product_difference_bank_fee.name,
                            'discount': False,
                            'tax_ids': False,
                            'load_ids': False,
                            'is_automatic_line': True,
                            'line_type': 'difference_bank_fee',
                            'account_id': product_difference_bank_fee.property_account_expense_id.id,
                        })]
                    })

    @api.depends('invoice_line_ids')
    def _compute_subtotal_loads(self):
        for record in self:
            discounts_totals = []
            record.subtotal_internal_loads = 0
            record.subtotal_internal_loads_discount = 0
            record.subtotal_internal_loads_additional = 0
            record.subtotal_internal_discounts = 0
            record.subtotal_internal_loads_subtotal = 0
            for line in record.invoice_line_ids:
                if line.is_automatic_line and not bool(line.benefit_discount_id) and not bool(line.unique_benefit_discounts) and line.product_id.barcode != '2SC':
                    record.subtotal_internal_loads_subtotal += line.price_subtotal
                    if line.line_type == 'Normal':
                        record.subtotal_internal_loads += line.price_subtotal
                    if line.line_type == '$ Descuento':
                        record.subtotal_internal_loads_discount += line.price_subtotal
                    if line.line_type == '$ Adicional':
                        record.subtotal_internal_loads_additional += line.price_subtotal
                if bool(line.benefit_discount_id) or bool(line.unique_benefit_discounts) or line.product_id.barcode == '2SC':
                    discounts_totals.append({'name': line.name, 'price_subtotal': line.price_subtotal})
                    record.subtotal_internal_discounts += line.price_subtotal
            record.discounts_totals = {'details': discounts_totals}

    def _compute_is_block_finished(self):
        for record in self:
            record.is_block_finished = not record.block_date_end or datetime.now() > record.block_date_end

    @api.depends('name')
    def _compute_draft_name(self):
        for record in self:
            if record.name == '/':
                formatted_date = record.create_date.strftime("%Y-%b-%d")
                if record.exo_invoice_sequence:
                    record.draft_name = f"{formatted_date}-{record.exo_invoice_sequence}"
                else:
                    record.draft_name = f"Borrador {formatted_date}"
            else:
                record.draft_name = record.name

    def _compute_payments(self):
        for move in self:
            if not move.invoice_payments_widget:
                continue
            widget_obj = json.loads(move.invoice_payments_widget)
            if not widget_obj or not widget_obj['content'] or len(widget_obj['content']) == 0:
                move.due_payment_time_days = False
                move.last_payment = False
                move.payment_time = 'Sin Fecha de Pago'
            else:
                move.last_payment = max([content['date'] for content in widget_obj['content']])
                move.payment_time = f'({(move.last_payment - move.invoice_date).days}) dias'
                term_days = 0
                for line in move.invoice_payment_term_id.line_ids:
                    if line.value == 'balance':
                        term_days += line.days
                move.due_payment_time_days = (move.last_payment - move.invoice_date).days - term_days

    def create_move_line_benefit_discount(self, current_env=None):
        current_env = current_env if current_env else self.env
        records = self.filtered(
            lambda f: f.is_automatic_invoice and not f.lock_invoice and f.state == 'draft' and 'in' in f.move_type
        )
        for move in records:
            move._apply_pending_recurrent_benefits(current_env)
            unique_benefits_to_create = move.partner_id.get_unique_benefits_line_to_create_to_this_months(current_env)
            if unique_benefits_to_create:
                self.create_unique_benefits_move_lines(move, unique_benefits_to_create, current_env)

    def _apply_pending_recurrent_benefits(self, current_env):
        self.ensure_one()
        move = self
        pending_benefits = current_env['partner.benefit.pending'].search([
            ('partner_id', '=', move.partner_id.id),
            ('state', '=', 'pending')
        ])
        if not pending_benefits:
            return
        new_lines_vals = []
        for pending in pending_benefits:
            vals = self._prepare_line_from_pending_benefit(pending)
            new_lines_vals.append((0, 0, vals))
        if new_lines_vals:
            move.sudo().write({'invoice_line_ids': new_lines_vals})
            newly_created_lines = move.invoice_line_ids.filtered(lambda l: l.pending_benefit_id.id in pending_benefits.ids)
            for line in newly_created_lines:
                try:
                    pending_rec = line.pending_benefit_id
                    if pending_rec:
                        pending_rec.write({
                            'state': 'applied',
                            'move_line_id': line.id,
                        })
                except Exception as e:
                    _logger.error(f"Error al actualizar el beneficio pendiente {line.pending_benefit_id.id}: {e}")

    def _prepare_line_from_pending_benefit(self, pending_benefit):
        self.ensure_one()
        product = pending_benefit.product_tmpl_id
        p_benefit_config = pending_benefit.partner_benefit_discount_id
        price_unit = pending_benefit.amount
        return {
            'pending_benefit_id': pending_benefit.id,
            'product_id': product.id,
            'quantity': pending_benefit.quantity,
            'analytic_account_id': p_benefit_config.analytic_account_id.id if p_benefit_config.analytic_account_id else False,
            'analytic_tag_ids': [(6, 0, p_benefit_config.analytic_tag_ids.ids)],
            'name': f"{product.name} - {p_benefit_config.benefit_discount_id.name} (Generado: {pending_benefit.generation_date.strftime('%d/%m/%Y')})",
            'price_unit': -abs(price_unit),
            'account_id': product.property_account_expense_id.id,
            'is_automatic_line': True,
        }

    def get_move_line_from_line_to_create(self, move, line_to_create, current_env):
        benefit_discount_id = line_to_create['p_benefit_discount']['benefit_discount_id']
        product = benefit_discount_id['product_tmpl_id']
        analytic_account_id = line_to_create['p_benefit_discount']['analytic_account_id']
        analytic_tag_ids = line_to_create['p_benefit_discount']['analytic_tag_ids']
        line_qty = line_to_create['p_benefit_discount']['product_quantity']
        taxes_result = product.taxes_id.compute_all(product.list_price, move.currency_id, quantity=line_qty, product=product, partner=move.partner_id)
        price_with_taxes = taxes_result['total_included']
        dic_tax_ids = product.taxes_id.ids + []
        value = (0, 0, {
            'benefit_discount_id': benefit_discount_id['id'],
            'benefit_discount_cicle_id': line_to_create['cicle']['id'],
            'product_id': product['id'],
            'quantity': line_qty,
            'analytic_account_id': analytic_account_id.id if analytic_account_id else False,
            'analytic_tag_ids': [(4, tag.id) for tag in analytic_tag_ids] if analytic_tag_ids and len(analytic_tag_ids) > 0 else False,
            'name': product['name'] + ' - ' + benefit_discount_id['name'] + ' - ' + line_to_create['cicle']['name'],
            'discount': False,
            'price_unit': -1 * product['list_price'],
            'tax_ids': dic_tax_ids,
            'account_id': product['property_account_expense_id']['id'] if move.move_type == 'out_invoice' else product['property_account_income_id']['id'],
            'debit': price_with_taxes if move.move_type == 'out_invoice' else 0,
            'credit': 0 if move.move_type == 'out_invoice' else price_with_taxes,
            'is_automatic_line': True
        })
        return value

    def create_benefits_move_lines(self, move, benefits_line_to_create, current_env):
        new_lines = [self.get_move_line_from_line_to_create(move, line_to_create, current_env) for line_to_create in benefits_line_to_create]
        move.sudo().write({'invoice_line_ids': new_lines})

    def insurance_2_percent_recompute(self, moves=None, current_env=None):
        try:
            current_env = current_env or http.request.env
            product_tmpl_2SC = current_env['product.template'].sudo().search([('barcode', '=', '2SC')], limit=1)
            if not product_tmpl_2SC:
                raise ValidationError("El Producto Seguro de Carga no fue encontrado, revise si existe un producto con el codigo de barra (2SC)")
            if moves:
                moves = current_env['account.move'].sudo().search([('id', 'in', moves)])
            else:
                moves = self
            for move in moves:
                if not move.partner_id.insurance_percent or move.partner_id.insurance_percent <= 0:
                    continue
                move_line = move.invoice_line_ids.filtered(lambda m_line: m_line.product_id.barcode == product_tmpl_2SC.barcode)
                if move_line and len(move_line) > 1:
                    raise ValidationError("El descuento de carga existe mas de una vez en la factura")
                move_total_positive_price = sum(line.price_subtotal for line in move.invoice_line_ids if line.price_subtotal > 0)
                move_total_positive_price = move_total_positive_price * move.partner_id.insurance_percent
                move_line_id = move_line.id if move_line else 0
                link_new_record = 1 if move_line else 0
                dic_tax_ids = product_tmpl_2SC.taxes_id.ids + []
                info = {
                    'invoice_line_ids': [(link_new_record, move_line_id, {
                        'price_unit': -1 * move_total_positive_price,
                        'product_id': product_tmpl_2SC.id,
                        'quantity': 1,
                        'name': product_tmpl_2SC.name,
                        'discount': False,
                        'tax_ids': dic_tax_ids,
                        'account_id': product_tmpl_2SC.property_account_income_id.id,
                        'is_automatic_line': True,
                    })]
                }
                move.sudo().write(info)
        except Exception as ex:
            raise ValidationError(f"Problemas al generar el 2% de Seguro de CArga")

    def create_unique_benefits_move_lines(self, move, unique_benefits_line_to_create, current_env):
        for b_line_to_create in unique_benefits_line_to_create:
            if not b_line_to_create['product_tmpl_id']['property_account_income_id'] or not b_line_to_create['product_tmpl_id']['property_account_income_id']['id']:
                raise ValidationError(f"Error no se ha encontrado la cuenta de ingreso a donde se reflejara los cambios em el producto {b_line_to_create['product_tmpl_id']['name']}")
            product = b_line_to_create['product_tmpl_id']
            taxes_result = product.taxes_id.compute_all(product.list_price, move.currency_id, quantity=b_line_to_create['product_quantity'], product=product, partner=move.partner_id)
            price_with_taxes = taxes_result['total_included']
            dic_tax_ids = product.taxes_id.ids + []
            move.sudo().write({
                'invoice_line_ids': [(0, 0, {
                    'product_id': product['id'],
                    'quantity': b_line_to_create['product_quantity'],
                    'name': product['name'] + ' - ' + b_line_to_create['name'],
                    'discount': False,
                    'price_unit': b_line_to_create['amount'] * (-1 if b_line_to_create["transaction_type"] == "credit" else 1),
                    'tax_ids': dic_tax_ids,
                    'move_id': move.id,
                    'account_id': product['property_account_expense_id']['id'] if b_line_to_create["transaction_type"] == "credit" else product['property_account_income_id']['id'],
                    'debit': price_with_taxes if b_line_to_create["transaction_type"] == "debit" else 0,
                    'credit': price_with_taxes if b_line_to_create["transaction_type"] == "credit" else 0,
                    'is_automatic_line': True,
                    'analytic_account_id': b_line_to_create['analytic_account_id']['id'] if b_line_to_create['analytic_account_id'] else None,
                    'unique_benefit_discounts': [(4, b_line_to_create.id)]
                })]
            })

    def send_change_payment_status_to_exo(self, trigger_record, trigger_all_records):
        load_ids = trigger_all_records.mapped('invoice_line_ids.load_ids').filtered(lambda r: r)
        if load_ids:
            load_ids.update_status_in_exo()

    def attach_url_images(self, record_line_load, json_object):
        message = ''
        transporter_images_files = json_object.get('transporterLoadCertification', {}).get('file', None)
        message = self.get_message_image(record_line_load, message, transporter_images_files, "Transportista")
        shipper_images_files = json_object.get('customerLoadCertification', {}).get('file', None)
        message = self.get_message_image(record_line_load, message, shipper_images_files, "Shipper")
        if message != '':
            self.with_context(mail_post_autofollow=True).message_post(body=message, subtype_xmlid="mail.mt_comment", message_type="comment")

    def get_message_image(self, record_line_load, message, shipper_images_files, title):
        if shipper_images_files:
            message += f'Imagenes {title}: <br />'
            for key in shipper_images_files:
                message += f'<a target="_blank" rel="noreferrer noopener" href="{shipper_images_files[key]}">{record_line_load.load_number}</a><br />'
            message += "<br/><br/> <hr />"
        return message

    def generate_client_template_files(self):
        current_env = http.request.env
        self.attach_invoice(current_env, False)
        return {'type': 'ir.actions.client', 'tag': 'reload'}

    def generate_provider_template_files(self):
        current_env = http.request.env
        self.attach_invoice(current_env, True)
        return {'type': 'ir.actions.client', 'tag': 'reload'}

    def attach_invoice(self, current_env, create_trasporter):
        current_env = current_env if current_env else self.env
        json_current_loads = current_env['account.line.load'].sudo().search([
            ('account_move_id', '=', self.id),
            ('json_current_load', '!=', None)
        ])
        json_data = []
        templates_all = current_env['template.load.file.property'].sudo().search([
            ('apply_for_all_providers', '=', self.id) if create_trasporter else ('apply_for_all_shippers', '=', self.id)
        ])
        for record_line_load in json_current_loads:
            try:
                json_object = json.loads(record_line_load.json_current_load)
                self.attach_url_images(record_line_load, json_object)
                json_data.append(json_object)
            except json.JSONDecodeError as e:
                raise ValidationError(f'Error decodificando el JSON_CURRENT_LOAD. Move {record_line_load.account_move_id.id}. Load_id: {record_line_load.id}: {e}')
            partner_templates = record_line_load.account_move_id.partner_id.templates_load_files
            templates_all += partner_templates
        templates_set = set(templates_all)
        templates = list(templates_set)
        for template in templates:
            binary_buffer = template.get_template_file(False, json_data)
            current_date = datetime.now().strftime("%Y-%m-%d_%I_%M_%S_%p")
            extention = 'pdf' if template.type == 'pdf' else 'xlsx'
            mimetype = 'application/pdf' if template.type == 'pdf' else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            file_name = f"{template.name}__{current_date}.{extention}"
            attachment = self.env['ir.attachment'].with_context(mail_post_autofollow=True).create({
                'name': file_name,
                'datas': binary_buffer,
                'mimetype': mimetype,
                'res_model': 'account.move',
                'res_id': self.id,
                'public': not template.is_internal
            })
            if not template.is_internal:
                self.with_context(mail_post_autofollow=True).message_post(body="Template: " + str(template.name), subtype_xmlid="mail.mt_comment", message_type="comment", attachment_ids=[attachment.id])
        self.sudo().write({'templates_uploaded': True})

    def get_info(self):
        self.ensure_one()
        move = self
        if move.move_type not in ['out_invoice', 'in_invoice']:
            raise ValidationError(f"Invalid move type for invoice {move.name}")
        moves = move.invoice_line_ids.sorted(key=lambda l: (-l.sequence, l.date, l.move_name, -l.id), reverse=True)
        moves_to_serialize = []
        for line in moves:
            moves_to_serialize.append({
                'id': line.id,
                'name': line.name,
                'discount': line.discount,
                'product_uom_id': {
                    'id': line.product_uom_id.id,
                    'name': line.product_uom_id.name,
                } if line.product_uom_id else None,
                'taxes': ', '.join(map(lambda x: (x.description or x.name), line.tax_ids)),
                'quantity': line.quantity,
                'price_unit': line.price_unit,
                'price_subtotal': line.price_subtotal,
                'tax_ids': [{"id": tax.id, "name": tax.name} for tax in line.tax_ids],
            })
        tax_totals = json.loads(move.tax_totals_json)
        subtotals = tax_totals["subtotals"]
        taxable_base = next((item['amount'] for item in subtotals if item["name"] == "Base imponible"), 0)
        total_to_pay = taxable_base + move.subtotal_internal_discounts
        taxes = tax_totals.get('groups_by_subtotal', {}).get("Base imponible", [])
        return {
            'allow_to_approve_invoice': move.state == 'draft' and move.provider_state == 'pending' and move.block_date_end and move.is_block_finished and datetime.now().date() >= (move.block_date_end + timedelta(days=1)).date(),

            'id': move.id,
            'name': move.name,
            'type': move.move_type,
            "state": move.state,
            "currency": move.currency_id.symbol,
            'partner': {
                'id': move.partner_id.id,
                'name': move.partner_id.name,
                'vat': move.partner_id.vat,
                'contact_address_complete': move.partner_id.contact_address_complete,
            },
            "company": {
                'id': move.company_id.id,
                'name': move.company_id.name,
                'vat': move.company_id.vat,
                'contact_address_complete': move.company_id.partner_id.contact_address_complete,
            },
            'invoice_date_timestamp': int(datetime.combine(move.invoice_date, datetime.min.time()).timestamp() * 1000) if move.invoice_date else None,
            'invoice_payment_term_id': {
                'id': move.invoice_payment_term_id.id,
                'name': move.invoice_payment_term_id.name,
                'note': move.invoice_payment_term_id.note
            },
            'subtotal_internal_loads': move.subtotal_internal_loads, # Sumas de las carga
            'subtotal_internal_loads_discount': move.subtotal_internal_loads_discount, # Descuentos a las cargas
            'subtotal_internal_loads_additional': move.subtotal_internal_loads_additional,# Pago adicional a las Cargas 
            'subtotal_internal_loads_subtotal': move.subtotal_internal_loads_subtotal, # Subtotal después de pagos y descuentos de las cargas
            'subtotal_after_internal_discounts': move.subtotal_internal_loads_subtotal + move.subtotal_internal_discounts, # Subtotal después de descuentos internos
            'taxable_base': taxable_base, # Subtotal después de impuestos
            'total_to_pay': total_to_pay, # Total a pagar..
            'amount_residual': move.amount_residual, # Total después de pago(s)
            'discounts_totals': move.discounts_totals['details'], # discounts
            'taxes': taxes, # Impuestos
            'invoice_line_ids': moves_to_serialize,

        }
